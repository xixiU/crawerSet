# -*- coding: utf-8 -*-
"""
Created on Fri Aug 17 20:15:16 2018

@author: x

直播吧
"""
import requests


from openpyxl import Workbook
from openpyxl import load_workbook

from bs4 import BeautifulSoup
import urllib.request
import json
import pickle as pl
import os,time
from datetime import datetime
import sys
import datetime
import csv

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3394.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Host': 'www.zhibo8.cc',
    'Referer': 'https://www.zhibo8.cc/index.htm',
    'Cache-Control': 'max-age=0',
    'Upgrade-Insecure-Requests': '1'}

    
class zhibo8:
    def __init__(self):
        self.session = requests.Session()
           
    def get_info(self):
        soup=self.get_soup()
        #print(soup)
        all_rows=[]
        try:
            wb = Workbook()# create wb
            #ws1 = wb.create_sheet(title="比赛")
            ws1 = wb['Sheet']
            ws1.sheet_properties.tabColor = "1072BA"
            ws1.cell(row= 1,column=1).value='日期'
            ws1.cell(row= 1,column=2).value='时间'
            ws1.cell(row= 1,column=3).value='赛事'
            ws1.cell(row= 1,column=4).value='A'
            ws1.cell(row= 1,column=5).value='B'
            #all_rows.append(['日期','时间','赛事','A','B'])
            current_row = 1
            #tags=soup.find('div',attrs={'class':'schedule_container left'})#remove right
            tags=soup.find_all('div',attrs={'class':'box'})
            if len(tags)>=1:
                for box in tags:
                    

                    time_tag= box.find('div',attrs={'class':'titlebar'})
                    context_tag = box.find('div',attrs={'class':'content'})
                    info_tag = context_tag.find_all('li')
                    for info in info_tag:
                            current_row+=1
                            info_text = info.get_text() 
                            info_text = info_text.replace('-','')
                            result = filter(None,info_text.split(' '))
                            for index ,value in enumerate(list(result)[0:4]):
                                ws1.cell(row= current_row,column=1).value=time_tag.text# write date
                                ws1.cell(row= current_row,column=index+2).value=value 

            wb.save(filename = '{}.xlsx'.format(datetime.datetime.now().isoformat().replace(':','-')))
            #with open('{}.xlsx'.format(datetime.datetime.now().isoformat().replace(':','-')),'w') as f_w:
                #fw_csv = csv.writer(f_w)
                #fw_csv.writerows(all_rows)
            print('ok')
        except Exception as e:
        
            print(e)
      
    def get_soup(self,url='https://www.zhibo8.cc/index2.htm'):
#        info = self.session.get(question_url,headers=headers,verify=False)#
#        if info.status_code !=200:
#            print('error')
#            sys.exit()
##        print(response.info())
        try:
            #response = urllib.request.urlopen(url,data=bytes(json.dumps(headers), encoding="utf-8"))
            info = self.session.get(url,headers=headers)#,verify=False
            #print(info.content)
            soup=BeautifulSoup(info.content,'html.parser')#response.read()
            return soup
        except urllib.error.HTTPError as e:
            print(e.code)
            print('未获取到soup对象')
            sys.exit()
            

if __name__ =='__main__':
    my8 =zhibo8()
    my8.get_info()
